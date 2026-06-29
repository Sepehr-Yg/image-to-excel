"""
image_to_excel.py
-----------------
Converts any image into a colorful Excel mosaic.
Each cell in the workbook represents a block of pixels from the original image,
filled with the average RGB color of that block.

Requirements:
    pip install pillow openpyxl

Usage:
    image_to_excel("photo.jpg")
    image_to_excel("photo.jpg", "output.xlsx", block_size=10)
"""

from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def image_to_excel(image_path, output_path="output.xlsx", block_size=10):
    """
    Convert an image to a colored Excel workbook mosaic.

    The image is divided into blocks of (block_size x block_size) pixels.
    Each block's average RGB color is used to fill a single Excel cell,
    reconstructing the image as a grid of colored cells.

    Args:
        image_path  (str): Path to the input image file (JPEG, PNG, etc.).
        output_path (str): Path for the output Excel file. Defaults to "output.xlsx".
        block_size  (int): Size of each pixel block in pixels. Defaults to 10.

    Returns:
        None. Saves the workbook to output_path.

    Example:
        >>> image_to_excel("photo.jpg", "mosaic.xlsx", block_size=10)
    """

    img = Image.open(image_path).convert("RGB")
    w, h = img.size

    # Crop to the nearest multiple of block_size to avoid partial blocks at the edges
    img = img.crop((0, 0, (w // block_size) * block_size, (h // block_size) * block_size))

    cols, rows = img.width // block_size, img.height // block_size

    # Resize the image to grid dimensions using LANCZOS resampling.
    # This is the core trick: each output pixel becomes the weighted average
    # of its corresponding block, replacing 4 nested loops with a single call.
    img = img.resize((cols, rows), Image.LANCZOS)

    wb = Workbook()
    ws = wb.active
    pixels = img.load()

    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 2.5

    for r in range(1, rows + 1):
        ws.row_dimensions[r].height = 15

    for r in range(rows):
        for c in range(cols):
            r_, g_, b_ = pixels[c, r]

            # Convert RGB tuple to hexadecimal color string
            hex_color = f"{r_:02X}{g_:02X}{b_:02X}"

            cell = ws.cell(row=r + 1, column=c + 1)
            cell.fill = PatternFill("solid", start_color=hex_color)

    wb.save(output_path)
    print(f"Done! Grid: {cols}x{rows} → {output_path}")


if __name__ == "__main__":
    image_to_excel(
        image_path="photo.jpg",
        output_path="output.xlsx",
        block_size=10,
    )